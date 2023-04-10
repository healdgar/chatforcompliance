import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, cell
from roboface import write_answer
import re
import os

def is_question(sentence):
    # Define a pattern to match verbs followed by a question mark.  This pattern finds setnences aht contain interrogatory pronouns, second person posessives, "please", "provide", or "Describe" or "Name of"

    pattern = r'\b(what|where|when|why|how|which|who|whom|whose)\b.*|\byou(r)?\b.*|\?.*|\bplease\b.*|\b(provide|describe)\b.*|Name of '

    # Use regular expressions to search for the pattern in the sentence
    match = re.search(pattern, sentence, re.MULTILINE)

    # If the pattern is found, the sentence is a question
    if match:
        return True
    else:
        return False

def insert_new_column(df, col_index, worksheet):
    for col in range(df.shape[1] - 1, col_index - 1, -1):
        target_col = get_column_letter(col + 2)
        source_col = get_column_letter(col + 1)
        if source_col in df:
            for row in range(df.shape[0]):
                if not is_merged_cell(worksheet, row, col):
                    df.at[row, target_col] = df.iloc[row, col]
        else:
            df[target_col] = ""


def is_merged_cell(worksheet, row, col):
    cell_coordinate = cell.get_column_letter(col + 1) + str(row + 1)
    for merged_cell_range in worksheet.merged_cells.ranges:
        if cell_coordinate in merged_cell_range:
            return True
    return False

last_inserted_columns = {}

def is_cell_in_merged_range(worksheet, row, col):
    cell_coordinate = get_column_letter(col + 1) + str(row + 1)
    for merged_cell_range in worksheet.merged_cells.ranges:
        if cell_coordinate in merged_cell_range:
            return True
    return False


# Define the Excel file name

excel_filename = input('Please enter the filename of the Excel that you placed in the Excels folder: ')

# Construct the full file path
script_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(script_dir)
excels_dir = os.path.join(parent_dir, 'Excels')
excel_filepath = os.path.join(excels_dir, excel_filename)

# Load the Excel file
xlsx = pd.read_excel(excel_filepath, sheet_name=None, header=None)

# Load the workbook before the loop
book = load_workbook(excel_filepath)

# Loop through each sheet
for sheet_name, df in xlsx.items():
    worksheet = book[sheet_name]  # Add this line to define the worksheet variable

    # Identify cells containing questions
    questions = []
    for row in range(df.shape[0]):
        for col in range(df.shape[1]):
            if is_merged_cell(worksheet, row, col):
                print(f"Skipped merged cell at ({row}, {col})")
            elif pd.notna(df.iloc[row, col]) and is_question(df.iloc[row, col]):
                questions.append((row, col))


    print("{} identified questions at coordinates: {}".format(len(questions), questions))
    # Create a new column to store the answers
    last_col = df.shape[1] - 1
    new_col = get_column_letter(last_col + 2)
    df[new_col] = ""

    input("Hit enter to start generating answers...")

    # Loop through the questions and write the answers
    for row, col in questions:
        question = df.iloc[row, col]
        answer = write_answer(question)

        if col in last_inserted_columns:
            last_inserted_columns[col] += 1
        else:
            last_inserted_columns[col] = col + 1

        insert_new_column(df, last_inserted_columns[col], worksheet)
        df.at[row, get_column_letter(last_inserted_columns[col])] = answer

        # Update the Excel file with the answers
        book = load_workbook(excel_filepath)
        writer = pd.ExcelWriter(excel_filepath, engine="openpyxl")
        writer.book = book
        ws = book[sheet_name]

        # Save the DataFrame to the same sheet it was read from, skipping merged cells
        for row in range(df.shape[0]):
            for col in range(df.shape[1]):
                if not is_cell_in_merged_range(ws, row, col):
                    ws.cell(row=row + 1, column=col + 1, value=df.iat[row, col])

writer.save()


# Print completion message
print("All questions have been answered and the Excel file has been updated.")
print(f"{len(answer)} answers have been written to column {new_col}")
