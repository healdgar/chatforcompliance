from time import thread_time_ns
from tkinter import END
import nltk
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from datetime import datetime
import csv
import requests
import getpass
import glob
import os
import openpyxl



def write_answer(searchTerm):
    #parameters
    # Example data
    answers = ""
    timestamp = datetime.utcnow().isoformat()
    print('timestamp',timestamp)
    model = 'gpt-3.5-turbo'
    temperature = 0.3
    max_tokens = 250
    user = getpass.getuser()

    # Tokenize and preprocess the search term
    searchTokens = word_tokenize(searchTerm.lower())
    stop_words = set(stopwords.words('english'))
    searchTokens = [token for token in searchTokens if token not in stop_words]

    # Find the most recently modified file that matches the filename pattern
    pattern = 'Client Answer Library*'
    suffix = '.xlsx'
    files = glob.glob(pattern + suffix)
    if len(files) == 0:
        print('Error: no files found for pattern', pattern + suffix)
        return
    filename = max(files, key=os.path.getmtime)

    # Load the Excel file into a list of dictionaries
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    csv_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, cell_value in enumerate(row):
            row_dict[header[i]] = cell_value
        csv_list.append(row_dict)


    # Calculate the similarity of each row to the search term
    matches = []
    for row in csv_list:
        line = (row['Question'] or '') + ' ' + (row['Supporting Answer'] or '')
        lineTokens = word_tokenize(line.lower())
        lineTokens = [token for token in lineTokens if token not in stop_words]
        similarity = sum([1 for token in searchTokens if token in lineTokens])
        matches.append({'Row': row, 'Similarity': similarity})

    # Sort the matches by similarity
    matches = sorted(matches, key=lambda k: k['Similarity'], reverse=True)

    # Select the top 10 matches
    top10 = []
    for match in matches[:6]:
        row = match['Row']
        answer = (row['Answer'] or '') + ' ' + (row['Supporting Answer'] or '')
        top10.append(answer)

    # Create a list of dictionaries with the search term, top 10 matches, and the messages array
    messages = [
#       {"role": "system", "content": "You are Rimini Street and are answering questions in a questionnaire.  Be sure to emphasize, where relevant, that Rimini Street does not process any customer data except for the support communications and does not process data as a servce and discuss the positive aspects of Rimini Street's compliance first."},
        {"role": "system", "content": "You are Rimini Street and are responding to customer questions.  Do not use conversational lead ins. Do not repeat the Question posed by client business or write 'Response to customer question:' Emphasize where appropriate that Rimini Street does not process client data as as service, and only processes support communications"},
        {"role": "user", "content": f"{searchTerm}? (please respond drawing from these example questions and answers about Rimini Street:)"},
        {"role": "user", "content": "\n".join(top10)}
    ]
    tuningContext = messages[:2]
    print("Search term", searchTerm)
    print("top10 matches",top10)
    # Print the number of tokens in the messages variable
    tokens = sum(len(message['content'].split()) for message in messages)
    print(f"Number of tokens in messages: {tokens}")

    # Create a dictionary with the model, messages array, temperature and max_tokens
    data = {
        "model": model,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": max_tokens,
    }

    # Convert the dictionary to JSON
    import json
    jsonData = json.dumps(data)
    # Set the API endpoint URL
    apiUrl = 'https://api.openai.com/v1/chat/completions'

    # Set the request headers
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Bearer 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'
    }

    # Send the JSON data to the API using the requests library

    try:
        response = requests.post(apiUrl, headers=headers, data=jsonData).json()
        answers = response['choices'][0]['message']['content']
        answer_forcsv = answers.replace('\n', ' ')
        print('Responsive answer: ',answers)

        # Open CSV file for writing
        with open('data.csv', 'a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)

            if file.tell() == 0:
                # File is empty, write header row
                writer.writerow(['timestamp','user', 'baseGPTmodel', 'creativity', 'max response size', 'searchTerm', 'tuning context', 'top10QRA responses', 'tokens', 'answer'])

            # Write data row
            writer.writerow([timestamp, user, model, temperature, max_tokens, searchTerm, tuningContext, top10, tokens, answers])
            print(writer.writerow)
    except requests.exceptions.RequestException as e:
        with open('error.txt', 'a') as file:
            file.write(str(e))
            file.write("\n")
        print("API call failed. Error message written to error.txt.")

        
    return answers



searchTerm = input("Enter search term: ")
write_answer(searchTerm)

# Wait for user to hit any key before closing the script



input("Hit enter to exit...")