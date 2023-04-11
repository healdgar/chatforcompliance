import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from roboface import write_answer
import re
import os

# Define the Excel file name
excel_filename = input('Please enter the filename of the Excel that you placed in the Excels folder: ')

# Construct the full file path
script_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(script_dir)
excels_dir = os.path.join(parent_dir, 'Excels')
excel_filepath = os.path.join(excels_dir, excel_filename)

# Load the Excel file
xlsx = pd.read_excel(excel_filepath, sheet_name=None, header=None)

def fake_write_answers(row, col):
    return f"Answer for ({row}, {col})"


def is_question(sentence):
    # Define a pattern to match verbs followed by a question mark.  This pattern finds setnences aht contain interrogatory pronouns, second person posessives, "please", "provide", or "Describe" or "Name of"

    pattern = r'\b(what|where|when|why|how|which|who|whom|whose)\b.*|\byou(r)?\b.*|\?.*|\bplease\b.*|\b(provide|describe)\b.*|Name of '

    # Use regular expressions to search for the pattern in the sentence
    match = re.search(pattern, sentence, re.MULTILINE)

    # If the pattern is found, the sentence is a question
    if match:
        return True
    else:
        return False

def update_question_coordinates(df, question_coordinates, col):
    updated_coordinates = []
    for r, c in question_coordinates:
        if c > col:
            updated_coordinates.append((r, c + 1))
            df.iat[r, c + 1] = df.iat[r, c]
            df.iat[r, c] = ""
        else:
            updated_coordinates.append((r, c))
    return updated_coordinates

def insert_new_column(df, col, last_inserted_columns):
    if col not in last_inserted_columns:
        new_col_index = col + 1
        new_col_name = f"RB_Responses_{len(last_inserted_columns) + 1}"
        last_inserted_columns[col] = new_col_name

        # Insert a new empty column to the right of the question column
        df.insert(new_col_index, new_col_name, "")

        # Write the column name to the second row of the new column
        df.at[1, new_col_name] = new_col_name
    else:
        new_col_name = last_inserted_columns[col]

    return new_col_name




last_inserted_columns = {}



# Loop through each sheet
for sheet_name, df in xlsx.items():

    # Identify cells containing questions and question columns
    questions = []
    question_columns = set()
    for row in range(df.shape[0]):
        for col in range(df.shape[1]):
            col_name = df.columns[col]
            if col_name not in last_inserted_columns.values():
                if pd.notna(df.iloc[row, col]) and is_question(df.iloc[row, col]):
                    questions.append((row, col))
                    question_columns.add(col)
    print("{} identified questions at coordinates: {}".format(len(questions), questions))

    question_coordinates = questions  # Define question_coordinates before updating it

    # Insert new columns for each question column
    question_answer_columns = {}
    for i, col in enumerate(question_columns):
        # Update the column index based on the number of previously inserted columns
        col += i
        new_col_name = insert_new_column(df, col, last_inserted_columns)

        # Store the new_col_name for each question column
        question_answer_columns[col - i] = new_col_name

        # Update the question coordinates only for the columns after the current question column
        question_coordinates = update_question_coordinates(df, question_coordinates, col)




    print("last_inserted_columns:", last_inserted_columns)

    # Process the questions and write the answers to the Excel file
    for row, original_col in questions:
        question = df.iloc[row, original_col]
        if original_col in question_answer_columns:
            answer_col_name = question_answer_columns[original_col]
            answer_col = df.columns.get_loc(answer_col_name)
            answer = fake_write_answers(row, original_col)
            df.at[row, answer_col] = answer
        else:
            print(f"Error: column {original_col} not found in question_answer_columns")



    # Update the Excel file with the answers
    book = load_workbook(excel_filepath)
    writer = pd.ExcelWriter(excel_filepath, engine="openpyxl")
    writer.book = book

    # Save the DataFrame to the same sheet it was read from
    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
    writer.save()

# Print completion message
print("All questions have been answered and the Excel file has been updated.")
