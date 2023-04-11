from time import thread_time_ns
from tkinter import END
import nltk
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from datetime import datetime
import csv
import requests
import getpass
import glob
import os
import openpyxl
import configparser
import json
import sys


# Read config.ini file
config = configparser.ConfigParser()

# Get the directory containing the script or executable
current_dir = os.path.dirname(os.path.abspath(sys.executable if getattr(sys, 'frozen', False) else __file__))
# Get config directory path
config_dir = os.path.join(os.path.dirname(current_dir), 'config')

# Construct the config file path
config_file_path = os.path.join(os.path.dirname(current_dir), 'config', 'config.ini')

# Read the config file
config.read(config_file_path)

# Get parameters from the config file
temperature = float(config.get('parameters', 'temperature'))
max_tokens = int(config.get('parameters', 'max_tokens'))
model = config.get('parameters', 'model')
answer_library_prefix = config.get('filepaths', 'answer_library_prefix')
num_matches = int(config.get('parameters', 'num_matches'))  # UPDATED: Read number of matches from the config file

# Get config directory path
config_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'config')

def write_answer(searchTerm):
    #variables

    answers = ""
    timestamp = datetime.utcnow().isoformat()
    user = getpass.getuser()

    # Tokenize and preprocess the search term
    searchTokens = word_tokenize(searchTerm.lower())
    stop_words = set(stopwords.words('english'))
    searchTokens = [token for token in searchTokens if token not in stop_words]

    # Find the most recently modified file that matches the filename pattern
    pattern = answer_library_prefix + '*'
    suffix = '.xlsx'
    files = glob.glob(os.path.join(config_dir, pattern) + suffix)
    if len(files) == 0:
        print('Error: no files found for pattern', os.path.join(config_dir, pattern) + suffix)
        return
    filename = max(files, key=os.path.getmtime)
    print('using QRA file: ', filename)

    # Load the Excel file
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    csv_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, cell_value in enumerate(row):
            row_dict[header[i]] = cell_value
        csv_list.append(row_dict)


    # Calculate the similarity of each row to the search term
    matches = []
    for row in csv_list:
        line = (row['Question'] or '') + ' ' + (row['Supporting Answer'] or '')
        lineTokens = word_tokenize(line.lower())
        lineTokens = [token for token in lineTokens if token not in stop_words]
        similarity = sum([1 for token in searchTokens if token in lineTokens])
        matches.append({'Row': row, 'Similarity': similarity})

    # Sort the matches by similarity
    matches = sorted(matches, key=lambda k: k['Similarity'], reverse=True)

    # Select the top n matches, where n is specified in the config file
    top_n = []  # Changed variable name to top_n
    for match in matches[:num_matches]:  #Use num_matches from the config file
        row = match['Row']
        answer = (row['Answer'] or '') + ' ' + (row['Supporting Answer'] or '')
        answer = answer.replace('Rimini Street', 'ACME')  # Replace "Rimini Street" with "ACME"
        top_n.append(answer)

    # Create a list of dictionaries with the search term, top n matches, and the messages array

    system_message = json.loads(config.get('parameters', 'system_message'))
    user_message = json.loads(config.get('parameters', 'user_message'))
    user_message['content'] = user_message['content'].format(searchTerm=searchTerm)

    messages = [system_message, user_message, {"role": "user", "content": "\n".join(top_n)}]  # UPDATED: Use top_n

    tuningContext = messages[:2]
    print("CLIENT QUESTION:", searchTerm)
    print("QRA CONTEXT:", top_n)
    # Print the number of tokens in the messages variable
    tokens = sum(len(message['content'].split()) for message in messages)
    print(f"Number of tokens in messages: {tokens}")

    # Create a dictionary with the model, messages array, temperature and max_tokens
    data = {
        "model": model,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": max_tokens,
    }

    # Convert the dictionary to JSON
    jsonData = json.dumps(data)
    # Set the API endpoint URL
    apiUrl = 'https://api.openai.com/v1/chat/completions'

    # Set the request headers
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {config.get("auth", "openai_api_key")}'
    }


    # Send the JSON data to the API using the requests library
    try:
        response = requests.post(apiUrl, headers=headers, data=jsonData).json()
        answers = response['choices'][0]['message']['content']
        print('ANSWER: ', answers)

        # Open CSV file for writing
        csv_path = os.path.join(config_dir, 'data.csv')
        with open(csv_path, 'a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)

            if file.tell() == 0:
                # File is empty, write header row
                writer.writerow(['timestamp', 'user', 'baseGPTmodel', 'creativity', 'max response size', 'searchTerm', 'tuning context', 'top_n_QRA_responses', 'tokens', 'answer'])  # UPDATED: top_n_QRA_responses

            # Write data row
            writer.writerow([timestamp, user, model, temperature, max_tokens, searchTerm, tuningContext, top_n, tokens, answers])  # UPDATED: Use top_n

    except requests.exceptions.RequestException as e:
        with open('error.txt', 'a') as file:
            file.write(str(e))
            file.write("\n")
        print("API call failed. Error message written to error.txt.")

    return answers

#searchTerm = input("Enter search term: ")
#write_answer(searchTerm)

# Wait for user to hit any key before closing the script
