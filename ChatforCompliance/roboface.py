import openai
import sys
import csv
import getpass
import glob
import os
import time
import openpyxl
import configparser
import json
import pickle
import numpy as np
from openai.embeddings_utils import get_embedding
from sklearn.metrics.pairwise import cosine_similarity
import sklearn.metrics._pairwise_distances_reduction._datasets_pair
import sklearn.metrics._pairwise_distances_reduction._middle_term_computer
from tqdm import tqdm
from datetime import datetime
from itertools import islice

# Roboface version 3-24-2023: adds multi-lingual semantic QRA matching, similarity threshold setting, and better logging
# Determine if the script is running as a compiled executable or as a Python script
if getattr(sys, 'frozen', False):
    # Running as a compiled executable
    script_dir = os.path.dirname(sys.executable)
else:
    # Running as a regular Python script
    script_dir = os.path.dirname(os.path.abspath(__file__))

# Construct the paths to the config directory (one level above the script or executable)
config_dir = os.path.join(script_dir, '..', 'config')

# Read the configuration files
config = configparser.ConfigParser()
config.read(os.path.join(config_dir, 'config.ini'))
secrets = configparser.ConfigParser()
secrets.read(os.path.join(config_dir, 'secrets.ini'))

# Get the API key from the secrets.ini file
api_key = secrets.get('auth', 'api_key')  # Change 'auth' to the correct section name if needed

# Set the API key for the openai library
openai.api_key = api_key

#initialize prior questions and answers
prior_questions = []
prior_answers = []

# Get parameters from the config file
temperature = float(config.get('parameters', 'temperature'))
max_tokens = int(config.get('parameters', 'max_tokens'))
model = config.get('parameters', 'model')
answer_library_prefix = config.get('filepaths', 'answer_library_prefix')
num_matches = int(config.get('parameters', 'num_matches'))
context_length = int(config.get('parameters', 'context_length'))

# Get config directory path
config_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'config')


def generate_embeddings_file(csv_list, filename):
    print(f"Embedding text in Client Answer Library ({filename})")
    embeddings_list = []
    
    # Get the API key from the secrets.ini file
    api_key = secrets.get('auth', 'api_key')
    
    # Update the OpenAI package configuration with the API key
    openai.api_key = api_key

    # Get replacements from config
    replacements = {}
    if config.has_section('replacements'):
        replacements = dict(config.items('replacements'))
    
    for i, row in enumerate(csv_list):
        line = (row['Question'] or '') + ' ' + (row['Answer'] or '') + ' ' + (row['Supporting Answer'] or '')
        for old, new in replacements.items():
            line = line.replace(old, new)
        line_embeddings = get_embedding(line, engine='text-embedding-ada-002')
        embeddings_list.append(line_embeddings)
        print(f"Progress: {i + 1}/{len(csv_list)}")
    with open(filename, 'wb') as f:
        pickle.dump(embeddings_list, f)
    print("Embeddings file created.")


# Find the most recently modified file that matches the filename pattern
pattern = answer_library_prefix + '*'
suffix = '.xlsx'
files = glob.glob(os.path.join(config_dir, pattern) + suffix)
if len(files) == 0:
    print('Error: no files found for pattern', os.path.join(config_dir, pattern) + suffix)

filename = max(files, key=os.path.getmtime)
print('using QRA file: ', filename)

# Load the Excel file
wb = openpyxl.load_workbook(filename)
ws = wb.active

header = [cell.value for cell in ws[1]]
csv_list = []
for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    row_dict = {'ExcelRow': row_index}  # Store the Excel row number
    for i, cell_value in enumerate(row):
        row_dict[header[i]] = cell_value
    csv_list.append(row_dict)

# Check if embeddings file exists; if not, generate it
embeddings_filename = os.path.join(config_dir, f'embeddings_{os.path.basename(filename)}.pickle')
if not os.path.exists(embeddings_filename):
    generate_embeddings_file(csv_list, embeddings_filename)

# Load the embeddings file
with open(embeddings_filename, 'rb') as f:
    embeddings_list = pickle.load(f)

def get_parameters(timestamp, user, baseGPTmodel, creativity, max_tokens, searchTerm, similarity_threshold, matching_context, tuning_context, top_n_matches, tokens, answer, source):
    return {
        'timestamp': timestamp,
        'user': user,
        'baseGPTmodel': baseGPTmodel,
        'creativity': creativity,
        'max_tokens': max_tokens,
        'searchTerm': searchTerm,
        'similarity_threshold': similarity_threshold,
        'matching_context': matching_context,
        'tuning_context': tuning_context,
        'top_n_matches': top_n_matches,
        'tokens': tokens,
        'answer': answer,
        'source': source
    }

def count_tokens(text):
    return len(text.split())

def truncate_prior_context(prior_questions, max_tokens=context_length):
    truncated_prior_questions = []
    current_tokens = 0

    for question in reversed(prior_questions):
        question_tokens = count_tokens(question)
        if current_tokens + question_tokens <= max_tokens:
            truncated_prior_questions.insert(0, question)
            current_tokens += question_tokens
        else:
            break

    return truncated_prior_questions


def _get_similarity_matches(searchTerm, csv_list, embeddings_list):
    # Get embeddings for the search term
    engine = "text-embedding-ada-002"
    search_embeddings = get_embedding(searchTerm, engine=engine)

    # Load similarity threshold from the configuration file
    similarity_threshold = float(config.get('parameters', 'similarity_threshold'))
        
    # Calculate the similarity of each row to the search term using embeddings
    matches = []
    for i, row in enumerate(csv_list):
        line_embeddings = embeddings_list[i]
        if line_embeddings is not None:
            search_embeddings_array = np.array(search_embeddings).reshape(1, -1)
            line_embeddings_array = np.array(line_embeddings).reshape(1, -1)
            similarity = cosine_similarity(search_embeddings_array, line_embeddings_array)[0, 0]
            matches.append({'Row': row, 'Similarity': similarity})

    matches.sort(key=lambda x: x['Similarity'], reverse=True)
    filtered_matches = [match for match in matches if match['Similarity'] >= similarity_threshold]

    # Get replacements from config
    replacements = {}
    if config.has_section('replacements'):
        replacements = dict(config.items('replacements'))

    matching_context = []
    tuning_context = []
    
    if not filtered_matches:
        matching_context = [f"No relevant content meeting the similarity threshold {similarity_threshold} were found in QRA."]
        tuning_context = [f"We have no specific information relating to this query."]
        top_n_matches = []
    else:
        top_n_matches = filtered_matches[:num_matches]
        for i, match in enumerate(top_n_matches):
            row = match['Row']
            excel_row_number = row['ExcelRow']
            similarity_score = match['Similarity']
            full_text = " Example Question: " + (row['Question'] or '') + " Example Answer: " + (row['Answer'] or '') + (row['Supporting Answer'] or '')
            for old, new_val in replacements.items():
                full_text = full_text.replace(old, new_val)
            
            display_entry = f"CAL Excel Row {excel_row_number} (Similarity: {similarity_score:.4f}): {full_text}"
            matching_context.append(display_entry)
            tuning_context.append(full_text)
            
    return top_n_matches, matching_context, tuning_context


def _prepare_openai_messages(searchTerm, prior_questions, prior_answers, tuning_context):
    system_message = json.loads(config.get('parameters', 'system_message'))
    user_message = json.loads(config.get('parameters', 'user_message'))
    user_message['content'] = user_message['content'].format(searchTerm=searchTerm)

    prior_context_messages = []
    for i in range(len(prior_questions)):
        prior_context_messages.append({"role": "user", "content": "Prior question: " + prior_questions[i]})
        if i < len(prior_answers):
            prior_context_messages.append({"role": "assistant", "content": "Prior answer: " + prior_answers[i]})

    tuning_context_entries = []
    for text in tuning_context:
        tuning_context_entries.append({"role": "user", "content": text})

    full_context_messages = [system_message] + prior_context_messages + [user_message] + tuning_context_entries
    return full_context_messages


def write_answer(searchTerm, source="direct Roboface query"):
    #variables
    answers = ""
    timestamp = datetime.utcnow().isoformat()
    user = getpass.getuser()
    global prior_questions
    global prior_answers

    # Truncate prior context
    prior_questions = truncate_prior_context(prior_questions)
    prior_answers = truncate_prior_context(prior_answers) # Assuming similar truncation logic for answers

    top_n_matches, matching_context, tuning_context = _get_similarity_matches(searchTerm, csv_list, embeddings_list)
    
    full_context_messages = _prepare_openai_messages(searchTerm, prior_questions, prior_answers, tuning_context)

    print("FULL CONTEXT MESSAGES", full_context_messages)
    print("\033[31m" + "PRIOR CONTEXT:" + "\033[0m", [msg for msg in full_context_messages if "Prior" in msg["content"]])
    print("\033[31m" + "CLIENT QUESTION:" + "\033[0m", searchTerm)
    print("\033[31m" + "QRA CONTEXT:" + "\033[0m", matching_context)
    
    tokens = sum(len(message['content'].split()) for message in full_context_messages)
    print(f"Number of tokens in messages: {tokens}")

    max_retries = 3
    retry_delay = 5  # seconds

    for attempt in range(max_retries):
        try:
            response = openai.ChatCompletion.create(
                model=model,
                messages=full_context_messages,
                temperature=temperature,
                max_tokens=max_tokens
            )
            answers = response.choices[0].message['content']
            print('\032[32mANSWER:\033[0m', answers)
            
            # If successful, break out of the retry loop
            break 
        except (openai.RateLimitError, openai.APIConnectionError, openai.Timeout) as e:
            print(f"Attempt {attempt + 1} of {max_retries} failed with retriable error: {str(e)}")
            if attempt + 1 == max_retries:
                print("Max retries reached. API call failed.")
                with open('error.txt', 'a', encoding='utf-8') as file:
                    file.write(f"OpenAI API Error (Max Retries Reached for {type(e).__name__}): {str(e)}\n")
                answers = "" # Ensure answers is empty or an error message
            else:
                time.sleep(retry_delay)
        except openai.APIError as e: # Non-retriable API errors (e.g., server-side issues)
            with open('error.txt', 'a', encoding='utf-8') as file:
                file.write(f"OpenAI API Error: {str(e)}\n")
            print(f"OpenAI API Error (Non-retriable). Error message written to error.txt: {str(e)}")
            answers = "" 
            break # Do not retry for these errors
        except openai.AuthenticationError as e: # Authentication issues
            with open('error.txt', 'a', encoding='utf-8') as file:
                file.write(f"OpenAI Authentication Error: {str(e)}\n")
            print(f"OpenAI Authentication Error. Error message written to error.txt: {str(e)}")
            answers = ""
            break # Do not retry for these errors
        except Exception as e: # Other unexpected errors
            print("Error content:", e)
            with open('error.txt', 'a', encoding='utf-8') as file:
                file.write(f"Unexpected error: {str(e)}\n")
            print("Unexpected error. Error message written to error.txt.")
            answers = ""
            break # Do not retry for these errors
    
    # Proceed only if 'answers' is not empty (i.e., API call was successful)
    if answers:
        parameters = get_parameters(
            timestamp=timestamp,
            user=user,
            baseGPTmodel=model,
            creativity=temperature,
            max_tokens=max_tokens,
            searchTerm=searchTerm,
            similarity_threshold=similarity_threshold,
            matching_context=matching_context,
            tuning_context=tuning_context,
            top_n_matches=top_n_matches,
            tokens=tokens,
            answer=answers,
            source=source  # Add the source parameter here
        )

        # Open CSV file for writing
        csv_path = os.path.join(config_dir, 'data.csv')
        with open(csv_path, 'a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)

            if file.tell() == 0:
                # File is empty, write header row
                writer.writerow(['timestamp', 'user', 'baseGPTmodel', 'creativity', 'max response size', 'searchTerm', 'similarity_threshold', 'matching context','tuning context', 'tokens', 'answer', 'source'])  # Add 'source' to the header row

            # Write data row
            writer.writerow([timestamp, user, model, temperature, max_tokens, searchTerm, similarity_threshold, matching_context, tuning_context, tokens, answers, source])  # Add the source parameter when writing the row

    #Append the current question and answer to the respective lists
    prior_questions.append(searchTerm)
    prior_answers.append(answers)

    return answers

# only run the script if it is called directly
# ask follow-up question and close the script if there is no additional question?

if __name__ == '__main__':
    search_term = input("Enter search term: ")
    write_answer(search_term)

    while True:
        follow_up = input("Do you have any additional questions? (Y/N): ")
        if follow_up == 'Y' or follow_up == 'y':
            search_term = input("Enter search term: ")
            write_answer(search_term)
        else:
            break
